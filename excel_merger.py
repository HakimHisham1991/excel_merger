from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional, Any, Union
import io
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import xlrd
import openpyxl

# ============================================================================
# CONFIG — edit these to customize for other jobs
# ============================================================================

# Folder containing the source Excel files to merge
SOURCE_DIR = Path(r"C:\Users\Public\Documents\011_EXCEL_MERGER\AL10_PARTS_EXCEL")

# Final merged output file (full path incl. filename)
OUTPUT_FILE = Path(r"C:\Users\Public\Documents\011_EXCEL_MERGER\AL10_PARTS_EXCEL_MERGED.xlsx")

# Which sheet to read from each source file.
# Use an int (0-based index, e.g. 0 = first sheet) or a str (exact sheet name).
SHEET_SELECTOR: Union[int, str] = 0

# Text marker that identifies the row where data starts (inclusive)
APPENDIX_HEADER = "Tool No."

# Text marker that identifies the row where data stops (exclusive). If not
# found, data is read until the last row/column of the sheet.
STOP_MARKER = "CAM Programmer :"

# Column (1-based) where "Part No." / "OP" / "Filename" values start getting
# written on each merged row. Set to None to always append after the last
# existing column of that row (safest, avoids overwriting data of varying
# widths). Set to an int (e.g. 1) to force them into fixed columns instead,
# e.g. START_COL_PART_OP = 18 -> Part No. in column R, OP in column S,
# Filename in column T, and existing row data beyond that point is shifted
# right to make room.
START_COL_PART_OP: Optional[int] = 18  # column R

# Filename pattern is "<OP>----<PART_NUMBER>" (e.g. "OP10----ABC123").
# Change this separator if your filenames use a different delimiter.
FILENAME_SEPARATOR = "----"

# Hard cap on how many columns to read per row. Protects against source
# files where formatting was applied to entire columns (e.g. max_column
# reports 16384 — Excel's absolute limit — instead of the real ~10-20
# columns of actual data). Raise this only if a real source sheet
# legitimately has more data columns than this.
MAX_COLUMNS_TO_SCAN = 200

# Files/prefixes to always skip when scanning SOURCE_DIR.
# "~$" skips Excel's temp lock files. The output file's own name (and
# numbered variants) are skipped so re-running the script never re-ingests
# its own previous outputs.
SKIP_PREFIXES = ("~$", OUTPUT_FILE.stem)

# ============================================================================


@dataclass(frozen=True)
class MergeConfig:
    source_dir: Path
    output_file: Path
    sheet_selector: Union[int, str]
    appendix_header: str
    stop_marker: str
    start_col_part_op: Optional[int]
    filename_separator: str
    skip_prefixes: tuple[str, ...]
    max_columns_to_scan: int = 200


EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}


def parse_filename(stem: str, separator: str) -> tuple[str, str] | None:
    """Return (part_number, op) parsed from filename stem, or None if pattern not matched.

    Expected filename pattern: "<OP><separator><PART_NUMBER>" (e.g. "OP10----ABC123").
    """
    if separator not in stem:
        return None
    op, part = stem.split(separator, 1)
    op = op.strip()
    part = part.strip()
    if not op or not part:
        return None
    return part, op


def iter_excel_files(source_dir: Path, skip_prefixes: tuple[str, ...]) -> Iterable[Path]:
    for p in sorted(source_dir.iterdir()):
        if not p.is_file():
            continue
        # Skip if name starts with any skip prefix (handles base and _N variants)
        if any(p.stem.startswith(prefix) for prefix in skip_prefixes):
            continue
        if p.suffix.lower() not in EXCEL_SUFFIXES:
            continue
        yield p


def get_sheet(wb_in: Any, is_xls: bool, sheet_selector: Union[int, str]) -> Any:
    """Return the target sheet object given an int index or str name."""
    if is_xls:
        if isinstance(sheet_selector, int):
            return wb_in.sheet_by_index(sheet_selector)
        return wb_in.sheet_by_name(sheet_selector)
    else:
        if isinstance(sheet_selector, int):
            return wb_in.worksheets[sheet_selector]
        return wb_in[sheet_selector]


def find_text_in_sheet(ws: Any, text: str) -> Optional[int]:
    """Search for text in sheet (supports both openpyxl and xlrd)"""
    text_norm = text.strip().lower()
    text_no_space = text_norm.replace(" ", "")

    if isinstance(ws, openpyxl.worksheet.worksheet.Worksheet):  # .xlsx / .xlsm
        for r in range(1, min(ws.max_row + 1, 300)):
            for c in range(1, min(ws.max_column + 1, 50)):
                val = ws.cell(row=r, column=c).value
                if not val:
                    continue
                val_str = str(val).strip().lower()
                if text_norm in val_str or text_no_space in val_str.replace(" ", ""):
                    return r
    else:  # xlrd (.xls)
        for r in range(ws.nrows):
            for c in range(ws.ncols):
                val = ws.cell_value(r, c)
                if not val:
                    continue
                val_str = str(val).strip().lower()
                if text_norm in val_str or text_no_space in val_str.replace(" ", ""):
                    return r + 1
    return None


MAX_EXCEL_CELL_LEN = 32767  # Excel's hard limit; longer strings corrupt the file


def sanitize_value(value: Any) -> Any:
    """Make a cell value safe to write into an .xlsx without corrupting it.

    - Strips illegal XML control characters (NUL, etc.) that Excel/openpyxl
      cannot serialize, which is the #1 cause of "we found a problem with
      some content" repair prompts.
    - Truncates strings over Excel's 32,767 character cell limit.
    - Converts NaN/Infinity floats (which are not valid XML numbers) to None.
    """
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if len(value) > MAX_EXCEL_CELL_LEN:
            value = value[:MAX_EXCEL_CELL_LEN]
        return value
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def sanitize_row(row_values: list) -> list:
    return [sanitize_value(v) for v in row_values]


def _try_build_workbook(rows: list[list]) -> tuple[bool, str]:
    """Build a throwaway workbook from `rows` and verify it survives a real
    save + reload round-trip (this is what actually catches corruption —
    many bad values don't raise a Python exception on write, they just
    produce invalid XML that only fails when something tries to parse it).
    Also catches values that DO raise while being appended (e.g. types
    openpyxl can't represent at all)."""
    wb = Workbook()
    ws = wb.active
    try:
        for row in rows:
            ws.append(row)
    except Exception as e:
        return False, f"append() raised: {e}"
    buf = io.BytesIO()
    try:
        wb.save(buf)
    except Exception as e:
        return False, f"save() raised: {e}"
    buf.seek(0)
    try:
        openpyxl.load_workbook(buf)
    except Exception as e:
        return False, f"reload raised: {e}"
    return True, ""


def _bisect_first_bad_row(known_good_rows: list[list], candidate_rows: list[list]) -> int:
    """Binary-search `candidate_rows` (appended after known_good_rows) for the
    first index whose inclusion breaks the round-trip. Returns that 0-based
    index. Assumes the caller already confirmed the full candidate list fails."""
    lo, hi = 0, len(candidate_rows) - 1
    bad_index = len(candidate_rows) - 1
    while lo <= hi:
        mid = (lo + hi) // 2
        ok, _ = _try_build_workbook(known_good_rows + candidate_rows[: mid + 1])
        if ok:
            lo = mid + 1
        else:
            bad_index = mid
            hi = mid - 1
    return bad_index


MAX_BISECTION_ROUNDS = 50  # safety cap in case many rows are bad


def insert_values(row_values: list, values: list, start_col: Optional[int]) -> list:
    """Place `values` (e.g. [part_number, op, filename]) into row_values either
    appended at the end (start_col=None) or inserted at a fixed 1-based column
    position, shifting existing values right."""
    if start_col is None:
        row_values.extend(values)
        return row_values

    idx = start_col - 1  # convert to 0-based
    # Pad with None if row is shorter than the insert point
    while len(row_values) < idx:
        row_values.append(None)
    row_values[idx:idx] = values
    return row_values


def get_unique_output_path(output_path: Path) -> Path:
    """Return a unique filename by appending _N if the file already exists."""
    if not output_path.exists():
        return output_path

    stem = output_path.stem
    suffix = output_path.suffix
    parent = output_path.parent
    counter = 1

    while True:
        new_path = parent / f"{stem}_{counter}{suffix}"
        if not new_path.exists():
            return new_path
        counter += 1


def merge_files(cfg: MergeConfig) -> None:
    cfg.output_file.parent.mkdir(parents=True, exist_ok=True)

    # Auto-rename output if it already exists
    final_output = get_unique_output_path(cfg.output_file)
    if final_output != cfg.output_file:
        print(f"Output file already exists. Using: {final_output.name}")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "MERGED"

    files = list(iter_excel_files(cfg.source_dir, cfg.skip_prefixes))
    total = len(files)
    print(f"Found {total} files to process.\n")

    merged_files = 0
    skipped = 0
    # Each entry: (source_filename, source_row_number, row_values)
    collected: list[tuple[str, int, list]] = []

    for idx, src_path in enumerate(files, start=1):
        print(f"[{idx}/{total}] {src_path.name}")

        # --- Derive Part No. / OP from filename ---
        parsed = parse_filename(src_path.stem, cfg.filename_separator)
        if parsed is None:
            print(f" ⚠️  Filename doesn't match '<OP>{cfg.filename_separator}<PART_NUMBER>' "
                  f"pattern; Part No./OP columns will be left blank.")
            part_number, op = "", ""
        else:
            part_number, op = parsed

        try:
            is_xls = src_path.suffix.lower() == ".xls"
            if is_xls:
                wb_in = xlrd.open_workbook(src_path)
            else:
                wb_in = openpyxl.load_workbook(src_path, data_only=False)

            ws = get_sheet(wb_in, is_xls, cfg.sheet_selector)
            sheet_label = ws.name if is_xls else ws.title
            print(f" ✅ Using sheet: '{sheet_label}'")

            # Find start row
            start_row = find_text_in_sheet(ws, cfg.appendix_header)
            if not start_row:
                print(f" ❌ '{cfg.appendix_header}' not found in sheet")
                skipped += 1
                continue

            print(f" Found '{cfg.appendix_header}' at row {start_row}")

            # Find stop row
            stop_row = find_text_in_sheet(ws, cfg.stop_marker)
            if is_xls:
                end_row = (stop_row - 1) if stop_row else ws.nrows
            else:
                end_row = (stop_row - 1) if stop_row else ws.max_row

            # Collect rows (not written to output yet — validated first)
            rows_merged = 0
            if is_xls:
                col_limit = min(ws.ncols, cfg.max_columns_to_scan)
                for r in range(start_row - 1, end_row):
                    row_values = [ws.cell_value(r, c) for c in range(col_limit)]
                    row_values = insert_values(
                        row_values, [part_number, op, src_path.name], cfg.start_col_part_op
                    )
                    collected.append((src_path.name, r + 1, sanitize_row(row_values)))
                    rows_merged += 1
            else:
                col_limit = min(ws.max_column, cfg.max_columns_to_scan)
                for r in range(start_row, end_row + 1):
                    row_values = [ws.cell(row=r, column=c).value for c in range(1, col_limit + 1)]
                    row_values = insert_values(
                        row_values, [part_number, op, src_path.name], cfg.start_col_part_op
                    )
                    collected.append((src_path.name, r, sanitize_row(row_values)))
                    rows_merged += 1

            merged_files += 1
            print(f" ✅ Collected {rows_merged} rows "
                  f"(Part No.='{part_number}', OP='{op}', Filename='{src_path.name}')\n")

        except Exception as e:
            print(f" ❌ Error: {e}\n")
            skipped += 1

    # --- Validate the full collected data set with a real save+reload round-trip ---
    print("Validating merged data (save + reload round-trip)...")
    all_row_values = [item[2] for item in collected]
    ok, err = _try_build_workbook(all_row_values)

    bad_rows_report: list[tuple[str, int, str]] = []

    if ok:
        good_rows = all_row_values
        print(" ✅ No corruption detected.\n")
    else:
        print(f" 🛑 Corruption detected: {err}")
        print(" Bisecting to find the exact offending row(s)...\n")
        good_rows: list[list] = []
        remaining = collected[:]  # list of (filename, source_row, row_values)
        rounds = 0

        while remaining and rounds < MAX_BISECTION_ROUNDS:
            rounds += 1
            candidate_values = [r[2] for r in remaining]
            ok2, err2 = _try_build_workbook(good_rows + candidate_values)
            if ok2:
                good_rows.extend(candidate_values)
                remaining = []
                break

            bad_idx = _bisect_first_bad_row(good_rows, candidate_values)
            culprit_file, culprit_src_row, _culprit_values = remaining[bad_idx]

            print(f" 🛑 CORRUPT ROW FOUND -> file='{culprit_file}', "
                  f"source row={culprit_src_row}  ({err2})")
            bad_rows_report.append((culprit_file, culprit_src_row, err2))

            # Everything before the culprit is confirmed safe; drop the
            # culprit itself; keep checking the rest for further bad rows.
            good_rows.extend(candidate_values[:bad_idx])
            remaining = remaining[bad_idx + 1:]

        if rounds >= MAX_BISECTION_ROUNDS and remaining:
            print(f" ⚠️  Hit bisection safety cap ({MAX_BISECTION_ROUNDS} rounds) "
                  f"with {len(remaining)} row(s) still unchecked. Adding them as-is.")
            good_rows.extend(r[2] for r in remaining)

    for row in good_rows:
        ws_out.append(row)

    wb_out.save(final_output)

    print("=" * 60)
    print(f"DONE! Files merged: {merged_files} | Files skipped: {skipped}")
    if bad_rows_report:
        print(f"Corrupt rows excluded from output ({len(bad_rows_report)}):")
        for fname, src_row, reason in bad_rows_report:
            print(f"  - {fname}  (source row {src_row})  [{reason}]")
    print(f"Output: {final_output}")


def main() -> None:
    cfg = MergeConfig(
        source_dir=SOURCE_DIR,
        output_file=OUTPUT_FILE,
        sheet_selector=SHEET_SELECTOR,
        appendix_header=APPENDIX_HEADER,
        stop_marker=STOP_MARKER,
        start_col_part_op=START_COL_PART_OP,
        filename_separator=FILENAME_SEPARATOR,
        skip_prefixes=SKIP_PREFIXES,
        max_columns_to_scan=MAX_COLUMNS_TO_SCAN,
    )
    merge_files(cfg)


if __name__ == "__main__":
    main()